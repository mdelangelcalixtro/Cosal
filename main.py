import openpyxl
import math
import time

#*****************Replit*******************************
#import inspect
#import os

#ruta_actual = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))

#inputx = 'input.xlsx'

#excel_file = os.path.join(ruta_actual, inputx)
#************************Replit*************************

#************************Colab**************************
#from google.colab import files

#uploaded = files.upload()

#excel_file = 'input.xlsx'
#************************Colab**************************

#************************Linux**************************
excel_file = '/home/miguel/Documents/Pycharm/Practices/test005/Cosal/input.xlsx'
#************************Linux**************************

#************************Linux**************************
#excel_file = 'C:/Users/Asus/Documents/Del Angel Calixtro Miguel Angel/input.xlsx'
#************************Linux**************************

#************************Android**************************
#excel_file = '/storage/emulated/0/Documents/Excel/input.xlsx'
#************************Android**************************

workbook = openpyxl.load_workbook(excel_file)

worksheet = workbook.active
def display_menu():
    print("Menu:")
    print("Presionar 1: Solicitud con medidas iguales")
    print("Presionar 2: Solicitud con medidas diferentes")
    print("Presionar 3: Salir del programa")

def get_user_choice():
    choice = input("Elige la opcion que desees: ")
    return choice

def main():
    while True:
        display_menu()
        user_choice = get_user_choice()

        if user_choice == '1':

            def calcula_espesor(D_1, d_1, mL_1):
                D_1 = D_1 * 10
                d_1 = d_1 * 10
                mL_1 = mL_1 * 1000
                espr = (math.pi * (D_1 ** 2 - d_1 ** 2)) / (4 * mL_1)
                return espr

            def calcula_metros_lineales_nuevos(espr, D_2, d_2):
                D_2 = D_2 / 100
                d_2 = d_2 / 100
                espr = espr / 1000
                mL_2 = (math.pi * (D_2 ** 2 - d_2 ** 2)) / (4 * espr)
                return mL_2

            def unidades_obtenidas_al_desembobinar(mL_1, mL_2):
                uod = mL_1 / mL_2
                return uod

            def unidades_obtenidas_al_dividir_por_ancho(ach_1, ach_2):
                uoda = ach_1 / ach_2
                return uoda

            def calcular_volumen(D_2, d_2, ach_2):
                rd = D_2 / 2
                gsr = D_2 - d_2
                vol = 2 * (math.pi) * rd * ach_2 * gsr
                return vol

            def calcular_peso_nuevo_por_unidad(ms, mL_2, ach_2, clb, vol):
                mL_2 = mL_2 * 100
                clb = clb / 10
                gme = ms / (mL_2 * ach_2)
                dsad = gme / clb
                uw_2 = (dsad * vol) / 1000
                return uw_2, gme, dsad

            def calcular_desperdicio(uw_2, uoda, uod, uw_1):
                smunid = int(uw_2) * int(uoda) * int(uod)
                desp = uw_1 - smunid
                return desp


            ms = float(input("Gramaje (g): "))
            D_1 = float(input("Diametro original exterior (cm): "))
            d_1 = float(input("Diametro original interior o Core (in): "))
            d_1 = d_1*2.54
            uw_1 = float(input("Peso total del rollo (kg): "))
            clb = float(input("Puntos (ppt): "))
            clb = clb*0.0025400
            mL_1 = float(input("Metros lineales (m): "))
            ach_1 = float(input("Medida original (cm):"))
            time.sleep(2)
            print("Datos que solicita el cliente")
            D_2 = float(input("Diametro solicitado exterior (cm): "))
            d_2 = float(input("Diametro solicitado interior o Core (in): "))
            d_2 = d_2*2.54
            ach_2 = float(input("Medida que solicita (cm): "))

            espr = calcula_espesor(D_1, d_1, mL_1)
            mL_2 = calcula_metros_lineales_nuevos(espr, D_2, d_2)
            uod = unidades_obtenidas_al_desembobinar(mL_1, mL_2)
            uoda = unidades_obtenidas_al_dividir_por_ancho(ach_1, ach_2)
            vol = calcular_volumen(D_2, d_2, ach_2)
            uw_2, gme, dsad = calcular_peso_nuevo_por_unidad(ms, mL_2, ach_2, clb, vol)
            desp = calcular_desperdicio(uw_2, uoda, uod, uw_1)

            #print("El espesor que se usa en ambos procesos es: ", espr, "mm")
            #print("Los metros lineales de cada nuevo rollo son: ", mL_2, "m")
            print("Bajadas: ", int(uod))
            print("Medidas obtenidas: ", int(uoda))
            #print("Volumen del rollo obtenido al final: ", vol, "cm3")
            print("Peso por rollo: ", uw_2, "kg")
            #print("Gramaje por peso entre area nueva: ", gme, "g/cm2")
            #print("Densidad de papel por gramaje entre calibre: ", dsad, "g/cm3")
            #print("Desperdicio de papel por recortes: ", desp, "kg")
            datos = ['Bajadas', 'Medidas obtenidas','Peso por rollo' ], [int(uod), int(uoda), uw_2], [' ', ' ', 'kg']
            for fila, datos_fila in enumerate(datos, start=1):
                for columna, valor in enumerate(datos_fila, start=1):
                    celda = worksheet.cell(row=columna, column=fila)
                    celda.value = valor

            workbook.save(excel_file)

#***************************************************************************************************************

        elif user_choice == '2':

            ms = float(input("Gramaje (g): "))
            D_1 = float(input("Diametro original exterior (cm): "))
            d_1 = float(input("Diametro original interior o Core (in): "))
            d_1 = d_1 * 2.54
            uw_1 = float(input("Peso total del rollo (kg): "))
            clb = float(input("Puntos (ppt): "))
            clb = clb * 0.0025400
            mL_1 = float(input("Metros lineales (m): "))
            time.sleep(2)
            print("Datos que solicita el cliente")
            numcort = int(input("Numero de medidas que quieres obtener: "))
            D_2 = float(input("Diametro solicitado exterior (cm): "))
            d_2 = float(input("Diametro solicitado interior o Core (in): "))
            d_2 = d_2 * 2.54

            #def calcula_espesor(D_1, d_1, mL_1):
            D_1 = D_1 * 10
            d_1 = d_1 * 10
            mL_4 = mL_1 * 1000
            espr1 = (math.pi * (D_1 ** 2 - d_1 ** 2)) / (4 * mL_4)
            #return espr
            #def calcula_metros_lineales_nuevos(espr, D_2, d_2):
            D_2 = D_2 / 100
            d_2 = d_2 / 100
            espr2 = espr1 / 1000
            mL_2 = (math.pi * (D_2 ** 2 - d_2 ** 2)) / (4 * espr2)
            #print("mL_2",mL_2)
            #return mL_2
            #def unidades_obtenidas_al_desembobinar(mL_1, mL_2):
            uod = mL_1 / mL_2
            #return uod
            #def unidades_obtenidas_al_dividir_por_ancho(ach_1, achs):
            #uoda = ach_1 / achs
            #return uoda
            #def calcular_peso_nuevo_por_unidad(ms, mL_2, clb, D_2,d_2,numcort, ttlcort):
            rd = D_2 / 2
            gsr = D_2 - d_2
            mL_3 = mL_2 * 100
            clb = clb / 10
            #promden = dasdtotal/numcort
            for i in range(numcort):
                achs = float(input(f"Ingresa medida {i + 1}: "))
                #ttlcort = 0
                dasdtotal = 0
                for i in range(numcort):
                    #Se requiera el doble for para sacar la sumatoria de las densidades
                    #achs = float(input(f"Ingresa medida {i + 1}: "))
                    gme = ms / (mL_3 * achs)
                    dsad = gme / clb
                    dasdtotal += dsad
                    #ttlcort += achs

            promden=dasdtotal/numcort
            # Lista para guardar valores vol y uw_2  para cada iteracion
            vol_values = []
            uw_2_values = []

            # Lazo cerrado para guardar datos en excel
            ttlcort = 0
            for i in range(numcort):
                print("Confirma medidas")
                achsx = float(input(f"Medida {i + 1}: "))
                vol = 2 * math.pi * rd * achsx * gsr
                #print(f"Volumen {i + 1}: ", vol)
                #vol_values.append(vol)  Guarda el dato en cada iteracion
                uw_2 = promden * vol * 10
                print(f"Peso {i + 1}: ", uw_2)
                uw_2_values.append(uw_2)  #Guarda el dato en cada iteracion
                ttlcort += achsx

                #Escribe vol y uw_2 a excel
                for i in range(numcort):
                    data = [['Pesos']]
                    for uw_2 in zip(uw_2_values):
                        data.append(uw_2)

                #Escribe todo al final en excel
                for fila, datos_fila in enumerate(data, start=1):
                    for columna, valor in enumerate(datos_fila, start=4):
                        celda = worksheet.cell(row=fila, column=columna)
                        celda.value = valor


            #print("Suma de densidades:", dasdtotal)
            #print("Promedio densidades:", promden)
            #print("El ancho total es:", ttlcort)
            #print("El espesor que se usa en ambos procesos es: ", espr1, "mm")
            #print("Los metros lineales de cada nuevo rollo son: ", mL_2, "m")
            print("Bajadas: ", int(uod))
            datos = ['Bajadas' ], [int(uod)]
            for fila, datos_fila in enumerate(datos, start=1):
                for columna, valor in enumerate(datos_fila, start=1):
                    celda = worksheet.cell(row=columna, column=fila)
                    celda.value = valor

            workbook.save(excel_file)


#***************************************************************************************************************

        elif user_choice == '3':
            print("\nSaliendo del programa. Adios!")
            break

        else:
            print("Numero invalido, elige la opcion del (1-3).")

        continue_or_not = input("\nDeseas continuar? (si/no): ")
        if continue_or_not.lower() != 'si':
            print("Saliendo del programa. Adios!")
            break

if __name__ == "__main__":
    main()